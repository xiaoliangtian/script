#!/usr/bin/env python
# coding=utf-8
# pylint: disable=import-error

import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from Config import DIR_SET

#inputFile = sys.argv[1]
geneFile = sys.argv[1]
#cnvDir = sys.argv[3]
#resDir = os.path.dirname(cnvDir) + '/'
#bamDir = resDir + 'dedup_bam/'
omimFile = DIR_SET['OMIM_CN']
clingenFile = '/data/local/internal/clingen_cnv.txt'
specificGenes = ['EFNB1','EHMT1','MBD5','MECP2','MED13L','MEF2C','NIPBL','NSD1','PAK2','PAX6','PLP1','PTCH1','PTHLH','SHH','TBL1XR1']
headList = ['chromosome:start-end','gene_num', 'gene', 'omim_num', 'omim_en', 'omim_cn', 'omim_gene', 'pa_region', 'region_Name', 'HI_score', 'TS_score', 'HI/TS_gene_num', 'HI/TS_gene']

with open(DIR_SET['CHROM_BANDS'], 'r') as j:
    bandsDb = json.load(j)

def query_band(chrNum, bands):
    bandA = ''
    bandB = ''
    outStr = ''
    bands = [ int(x) for x in bands ]
    for key in bandsDb.keys():
        locs = key.split('_')
        if chrNum == locs[0] and len(locs) > 1:
            start = int(locs[1])
            end = int(locs[2])
            if start <= bands[0] <= end:
                bandA = locs[0].replace('chr', '') + bandsDb[key]
            if start <= bands[1] <= end:
                bandB = locs [0].replace('chr', '') + bandsDb[key]
            elif bands[1] == bandsDb[chrNum]:
                bandB = 'qter'

            if bandA and bandB:
                if bandA == bandB:
                    outStr = bandA
                else:
                    outStr = '-'.join([bandA, bandB])
                break
    return outStr

def access_omim(filename):
    items = {}
    with open(filename) as omim:
        omim.readline()
        for line in omim:
            line = line.strip()
            lines = line.split('\t')
            if len(lines) > 10:
                gene = lines[14]
                phenoStr = lines[12].replace('"', '')
                phenoCnStr = lines[13].replace('"', '')
                if gene not in items.keys():
                    items[gene] = {'en': phenoStr, 'cn': phenoCnStr}
    return items

def access_clingen(filename):
    items = {}
    keyGene = {}
    with open(filename) as clingen:
        clingen.readline()
        for line in clingen:
            line = line.strip().split('\t')
            gene = line[1].replace(' ', '').split(',')
            if gene[0] == '.':
                pos = line[0].strip()
                if pos not in items.keys():
                    items[pos] = line[0] + '\t' + '\t'.join(line[2:])

            for sub in gene:
                sub = sub.strip()
                if sub != '.' and sub not in keyGene.keys():
                    keyGene[sub] = line[0] + '\t' + '\t'.join(line[2:])

    return items, keyGene

def access_loc(filename):
    header = '\t'.join(headList)
    print(header)
    items = {}
    clingen_db, keyGenes = access_clingen(clingenFile)
    omim_db = access_omim(omimFile)
    with open(filename) as locFile:
        for line in locFile:
            line = line.strip().split('\t')
            loc = line[0].replace('..', '-')
            if loc not in items.keys():
                lineStr = line[1] + '\t' + line[2]
                genes = line[2].split('|')
                geneList = []
                geneEnList = []
                geneCnList = []
                clingenList = []
                spGeneList = []
                for gene in genes:
                    if gene in omim_db.keys():
                        geneList.append(gene)
                        genePheno = gene + '[' + omim_db[gene]['en'] + ']'
                        geneEnList.append(genePheno)
                        geneCnPheno = gene + '[' + omim_db[gene]['cn'] + ']'
                        geneCnList.append(geneCnPheno)
                    if gene in keyGenes.keys():
                        clingenList.append(keyGenes[gene])
                    if gene in specificGenes:
                        spGeneList.append(gene)

                if len(geneList) > 0:
                    geneStr = ','.join(geneList)
                else:
                    geneStr = '.'
                if len(geneEnList) > 0:
                    geneEnStr = ','.join(geneEnList)
                else:
                    geneEnStr = '.'
                if len(geneCnList) > 0:
                    geneCnStr = ','.join(geneCnList)
                else:
                    geneCnStr = '.'
                if len(spGeneList) > 0:
                    spGeneStr = ','.join(spGeneList)
                else:
                    spGeneStr = '.'

                clingenStr = '\t'.join(['.', '.', '.', '.'])
                '''
                if len(clingenList) > 0:
                    print (clingenList)
                    for Str in clingenList:
                        str=
                    clingenStr = '|'.join(clingenList)
                '''
                if len(clingenList) == 1:
                    clingenStr = clingenList[0]
                elif len(clingenList) > 1:
                    tmpStrA = []
                    tmpStrB = []
                    tmpStrC = []
                    tmpStrD = []
                    for item in clingenList:
                        tmpList = item.split('\t')
                        tmpStrA.append(tmpList[0])
                        tmpStrB.append(tmpList[1])
                        tmpStrC.append(tmpList[2])
                        tmpStrD.append(tmpList[3])
                    clingenStr = '\t'.join(['|'.join(tmpStrA), '|'.join(tmpStrB), '|'.join(tmpStrC), '|'.join(tmpStrD)])
                else:
                    posChr = loc.split(':')[0]
                    posRange = loc.split(':')[1].split('-')
                    posStart = int(posRange[0])
                    posEnd = int(posRange[1])
                    posLen = posEnd - posStart
                    posLimit = 0.5
                    if posLen < 100000:
                        posLimit = 0.8
                    for posItem in clingen_db.keys():
                        itemChr = posItem.split(':')[0]
                        itemRange = posItem.split(':')[1].split('-')
                        itemStart = int(itemRange[0])
                        itemEnd = int(itemRange[1])
                        if posChr == itemChr:
                            if ( posStart >= itemStart and posEnd <= itemEnd ) or ( posStart <= itemStart and posEnd >= itemEnd):
                                clingenStr = clingen_db[posItem]
                                break
                            elif ( posStart <= itemStart and posEnd <= itemEnd ) and (posEnd - itemStart)/posLen >= posLimit:
                                clingenStr = clingen_db[posItem]
                                break
                            elif ( posStart >= itemStart and posEnd >= itemEnd ) and (itemEnd - posStart)/posLen >= posLimit:
                                clingenStr = clingen_db[posItem]
                                break

                lineStr = '\t'.join([lineStr, str(len(geneEnList)), geneEnStr,  geneCnStr,  geneStr, clingenStr, str(len(spGeneList)), spGeneStr])
                items[loc] = lineStr
            print(loc +'\t'+ lineStr)
    return items
'''
def access_bams(path):
    temp_files = []
    for file in os.listdir(path):
        if '.dedup.sorted.bam' in file:
            filename = file.split('.')[0].split('_')[0]
            if filename not in temp_files:
                temp_files.append(filename)
    return temp_files

def sort_list(item):
    sort_item = item.split('\t')[2].split(':')[0].replace("chr", '')
    if sort_item == 'X':
        sort_key = 23
    elif sort_item == 'Y':
        sort_key = 24
    elif sort_item == 'MT':
        sort_key = 0
    else:
        sort_key = int(sort_item)
    return sort_key

def form_data(snpLine, heads):
    tmp_dict = {}
    title = heads.split('\t')
    snpInfo = snpLine.split('\t')
    for i in range(0, len(title)):
        try:
            snpInfo[i] = float(snpInfo[i])
        except:
            pass
        tmp_dict[title[i]] = snpInfo[i]
    return tmp_dict

def access_xlsx(sample, data):

    def highlight(data):
        rowData = pd.Series(data, index=data.index)
        if int(rowData['KB']) >= 1000:
            return pd.Series('background-color: {}'.format('#ffff00'), rowData.index)
        else:
            return pd.Series('', rowData.index)

    xlsxFile = resDir + sample + '.whanno.xlsx'
    if os.path.exists(xlsxFile):
        book = load_workbook(xlsxFile)
        df = pd.DataFrame(data, columns=data[0].keys())
        writer = pd.ExcelWriter(xlsxFile, engine='openpyxl')
        writer.book = book
        df.style.\
            apply(highlight, axis=1).\
            to_excel(writer,'CNV', index=False)
        worksheet = writer.sheets['CNV']
        FullRange = "A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
        worksheet.auto_filter.ref = FullRange
        writer.save()
        writer.close()
        print (xlsxFile, 'Done')
'''
res_db = {}
df_db = {}
loc_db = access_loc(geneFile)
#bam_files = access_bams(bamDir)
headList = ['band', 'gene_num', 'gene', 'omim_en', 'omim_cn', 'omim_gene', 'pa_region', 'region_Name', 'HI_score', 'TS_score', 'HI/TS_gene']
#print(bam_files)

'''
with open(inputFile) as inputs:
    header = inputs.readline().strip() + '\t' + '\t'.join(headList)
    for line in inputs:
        line = line.strip()
        lines = line.split('\t')
        loc = lines[2]
        chrNum = loc.split(':')[0]
        bands = loc.split(':')[1].split('-')
        bandStr = query_band(chrNum, bands)
        line += '\t' + bandStr
        sample = lines[0].split('.')[0].split('_')[0]
        if len(bam_files) == 0:
            if 'DATA' not in res_db.keys():
                res_db['DATA'] = []
            if loc in loc_db.keys():
                line += '\t' + loc_db[loc]
            res_db['DATA'].append(line)
        elif sample in bam_files:
            if sample not in res_db.keys():
                res_db[sample] = []
            if sample not in df_db.keys():
                df_db[sample] = []
            if loc in loc_db.keys():
                line += '\t' + loc_db[loc]
            res_db[sample].append(line)
            df_db[sample].append(form_data(line, header))

for sample in res_db.keys():
    if sample in df_db.keys():
        access_xlsx(sample, df_db[sample])
    res_db[sample].sort(key=sort_list)
    res_db[sample].insert(0, header)
    sample_res = '\n'.join(res_db[sample])
    cnvFile = cnvDir + '/' + sample + '.cnv.result'
    with open(cnvFile, 'w') as outfile:
        outfile.write(sample_res)
'''
