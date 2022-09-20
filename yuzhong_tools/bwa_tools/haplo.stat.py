#!/usr/bin/python

# coding=utf-8

import os,sys,argparse
from subprocess import Popen,PIPE
from openpyxl import load_workbook,Workbook

parser = argparse.ArgumentParser(description='Statistic haplotype by me',epilog='It is used for statistic haplotype analyse')

parser.add_argument('--version', action='version', version='%(prog)s 1.0')
parser.add_argument('-i','--input',required=True, help='input vcf files')
parser.add_argument('-t',metavar='F or M',required=True, help='analyse haplotype for from F or M(example:F)')
parser.add_argument('-o',default='test.xlsx', help='output file')
#parser.add_argument("--nochain",action = "store_true",default=False, help="no chain snp" )
args = parser.parse_args()

inputfile=args.input
outFile = args.o
#haplo=args.t

hap_script = '/home/tianxl/pipeline/yuzhong_tools/bwa_tools/get_si2_v6_ForM_poisson_v3_gatk.py'
son2Idx,dadIdx,momIdx,son1Idx = {},{},{},{}

def hap_ana(son2col,son1col,fcol,mcol,head=False):
    cmd = ['python', hap_script, '-i', inputfile, '-f', fcol, '-m', mcol, '-s1', son1col, '-s2', son2col, '-t', args.t, '-o1', 'test'] 
    
    p = Popen(cmd, stdout=PIPE, stderr=PIPE)
    p.wait()
    stdout, stderr = p.communicate()
    out = stdout.decode('utf-8')
    return out

#read vcf file
with open(inputfile,"r") as textfile:
    header = textfile.readline()
    heads = header.strip().split('\t')
    for head in heads:
        idx = heads.index(head)
        if head[0:2] == 'RD' and idx > 0:
            sample = head[3:].split('_')[0]
            #print(sample)
            if 'CF' in sample:
                son2Idx[sample] = str(idx)
            elif 'F' in sample:
                dadIdx[sample] = str(idx)
            elif 'M' in sample:
                momIdx[sample] = str(idx)
            elif 'YS' in sample or 'S2' in sample:
                son1Idx[sample] = str(idx)


haploOut = []
#haplotype analyse
for h in son2Idx.keys():
    h1 = h[2:]
    #print(h1)
    for i in son1Idx.keys():
        i1 = i[2:]
        for j in dadIdx.keys():
            j1 = j[1:]
            for k in momIdx.keys():
                k1 = k[1:]
                #print(h[-4:])
                if h1 == i1 == j1 == k1:
                    haplo = hap_ana(son2Idx[h],son1Idx[i],dadIdx[j],momIdx[k])
                    haploOut.append(haplo)

single = []
if os.path.exists(outFile):
    wb = load_workbook(outFile)
    ws = wb.create_sheet(args.t + 'hap')
else:
    wb = Workbook()
    ws = wb.active
    ws.title = args.t + 'hap'

for n in range(len(haploOut)):
    single = haploOut[n].split("\n")
    for m in range(len(single)):
        ws.append(single[m].split("\t")) 
wb.save(outFile)
