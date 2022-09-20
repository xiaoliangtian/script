#!/usr/local/python36/bin/python3
# -*- coding: utf-8 -*- 
import  xdrlib ,sys, os
import xlrd
import pandas as pd
import gzip
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook,Workbook
import numpy as np
from multiprocessing import Pool,Queue,Process

input = sys.argv[1]
inputList = input.split(',')
#打开excel文件
def open_excel(file):
    try:
        data = xlrd.open_workbook(file)
        return data
    except Exception as e:
        print (str(e))


#sfGenes = ['BRCA1','BRCA2','TP53','STK11','MLH1','MSH2','MSH6','PMS2','APC','MUTYH','VHL','MEN1','RET','PTEN','RB1','SDHD','SDHAF2','SDHC','SDHB','TSC1','TSC2','WT1','NF2','COL3A1','FBN1','TGFBR1','TGFBR2','SMAD3','ACTA2','MYH11','MYBPC3','MYH7','TNNT2','TNNI3','TPM1','MYL3','ACTC1','PRKAG2','GLA','MYL2','LMNA','RYR2','PKP2','DSP','DSC2','TMEM43','DSG2','KCNQ1','KCNH2','SCN5A','LDLR','APOB','PCSK9','RYR1','CACNA1S','BMPR1A','SMAD4','ATP7B','OTC']

sfGenes = []

def access_xlsx(sample, data, fileName):

    def highlight(data):
        for gene in sfGenes:
            if data == gene:
                return 'color: #DC143C'
        return ''
    
    def highlight_cnv(data):
        rowData = pd.Series(data, index=data.index)
        if int(rowData['KB']) >= 1000:
            return pd.Series('background-color: {}'.format('#ffff00'), rowData.index)
        else:
            return pd.Series('', rowData.index)
 
    xlsxFile = fileName
    
    
    if os.path.exists(xlsxFile):
        book = load_workbook(xlsxFile)
        df = pd.DataFrame(data, columns=data[0].keys())
        writer = pd.ExcelWriter(xlsxFile, engine='openpyxl')
        writer.book = book
        df.style.\
            to_excel(writer,sample, index=False)
        worksheet = writer.sheets[sample]
        FullRange = "A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
        worksheet.auto_filter.ref = FullRange
        writer.save()
        writer.close()
        print(xlsxFile, 'done!')
        
    
    else:
        
        df = pd.DataFrame(data, columns=data[0].keys())
        writer = pd.ExcelWriter(xlsxFile, engine='xlsxwriter')
        writer.book.use_zip64()
        
        if sample == 'standard':
            df.style.\
                to_excel(writer,sample, index=False)
        
        worksheet = writer.sheets[sample]
        worksheet.autofilter('A1:FA1')
        worksheet.freeze_panes(1, 0)
        writer.save()
        writer.close()

def form_data(snpLine, keys):
    tmp_dict = {}
    title = keys.split('\t')
    snpInfo = snpLine.split('\t')
    for i in range(0, len(title)):
        try:
            snpInfo[i] = float(snpInfo[i])
        except:
            pass
        tmp_dict[title[i]] = snpInfo[i]
    return tmp_dict

#根据名称获取Excel表格中的数据   参数:file：Excel文件路径     colnameindex：表头列名所在行的索引  ，by_name：Sheet1名称
def excel_table_byname(file):
    tableList = ['standard','CNV','MitoVars']
    list1 = list(range(126))
    output = file.split('.')[0]+'.filter.xlsx'
    if os.path.exists(output):
        os.system('rm ' + output)
    data = open_excel(file) 
    sheetnames = data.sheet_names()
    # print(sheetnames)
    for by_name in tableList:
        if by_name in sheetnames:
            colnameindex = 0
            table = data.sheet_by_name(by_name)
            nrows = table.nrows 
            ncols = table.ncols
            colnames = table.row_values(colnameindex) 
            #cols = '\t'.join(colnames)
            listn =[]
            list2 = []
            col = []
            if by_name == 'standard':
                # list1 = list(range(ncols))
                for index, n in enumerate(colnames):
                    if n not in standRmList:
                        col.append(n)
                        list2.append(index)
                # colnames = colnames[:88] + colnames[92:93]+colnames[99:100]+colnames[102:105]+colnames[106:109]+ colnames[111:115]
                # list2 = list1[:88] + list1[92:93]+list1[99:100]+list1[102:105]+list1[106:109]+list1[111:115]
            elif by_name == 'DECA_CNV':
                # list1 = list(range(ncols))
                for index, n in enumerate(colnames):
                    if n not in DECARmList:
                        col.append(n)
                        list2.append(index)
                # colnames = colnames[:6]+colnames[7:10]+colnames[11:13]
                # list2 = list1[:6]+list1[7:10]+list1[11:13]
            elif by_name == 'MitoVars':
                for index, n in enumerate(colnames):
                    if n not in MitoRmList:
                        col.append(n)
                        list2.append(index)
                # colnames = colnames
                # list2 = list(range(ncols))
            else:
                col = colnames[:4]+colnames[15:18]+colnames[19:21]
                list2 = list1[:4]+list1[15:18]+list1[19:21]
            cols = '\t'.join(col)
            #print(nrows)
            for rownum in range(1, nrows): 
                row = table.row_values(rownum) 
                if row: 
                    app = [] 
                    for i in list2:
                        app.append(str(row[i]))
                    newSnpLine = '\t'.join(app)
                    exlData = form_data(newSnpLine, cols)
                    listn.append(exlData) #装载数据
            access_xlsx(by_name, listn, output)
    

#主函数
def main():
    pool = Pool(processes = 4)
    pool.map(excel_table_byname, inputList)
    pool.close()
    pool.join()

standRmList = ['GeneDetail.refGene','genomicSuperDups','InterVar_automated','InterVar_proof','Otherinfo','DP','#CHROM','POS','ID','REF','ALT','INFO','FORMAT','gene_dis','pred_ratio','pred_stats','local_fre','fre_count','DS_AG','DS_AL','DS_DG','DS_DL','local_var','EYE_PANEL','ENDO_PANEL','META_PANEL','NEVR_PANEL','dup_region']
DECARmList = ['info','omim_en','pa_region','region_Name','HI_score','TS_score','HI/TS_gene']
MitoRmList = []
if __name__=="__main__":
    main()

